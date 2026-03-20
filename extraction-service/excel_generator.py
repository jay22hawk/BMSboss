"""
BMS Boss — Excel Generator
Populates the actual 2026 Prescriptive BMS Calculator template (.xlsx).

The template has 27+ interconnected sheets with formulas. We populate:
  1. "PA Only - Pre-Inspection" sheet: Company info, accounts (feeds into User Inputs via formulas)
  2. "User Inputs and Savings" sheet: Building energy data, project info, affected areas, sequences

Cell mapping is based on analysis of the actual 2026 V1.0 template.
Orange cells (theme=5, tint=0.4) = direct user input fields.
Blue cells (theme=8, tint=0.8) = fields that feed from Pre-Inspection via formulas.
"""

import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from models import BMSCalculatorInput, ExtractedBillData, GenerationResponse
from enum import Enum

TEMPLATE_FILENAME = "2026_Prescriptive_BMS_Calculator_V1_0.xlsx"

# Sheet names
SHEET_USER_INPUTS = "User Inputs and Savings"
SHEET_PRE_INSPECTION = "PA Only - Pre-Inspection"

# ─── Cell Mappings ────────────────────────────────────────────────────────────

# PA Only - Pre-Inspection sheet: these cells feed into User Inputs via formulas
PRE_INSPECTION_MAP = {
    "company_name": "C9",           # Company Name
    "company_address": "C10",       # Company Address (street portion)
    "company_city": "F10",          # City
    "customer_contact_name": "C11", # Customer Contact Name
    "customer_phone": "F11",        # Phone
    "pa_technical_rep": "C12",      # PA Technical Representative
    "pa_tech_rep_phone": "C13",     # PA Tech Rep Phone
    "application_number": "H13",    # Application #
    "electric_account": "C14",      # Electric accounts
    "gas_account": "C15",           # Gas accounts
    "electric_pa": "I11",           # Electric PA name
    "gas_pa": "I10",                # Gas PA name
}

# User Inputs and Savings: direct input cells (orange, not formula-linked)
USER_INPUTS_MAP = {
    # Building Energy Use Intensity (rows 17-21)
    "building_activity": "C17",     # Dropdown: building activity
    "heating_fuel": "C18",          # Dropdown: heating fuel
    "total_building_sqft": "C19",   # Total Building Area (sqft)
    "annual_electric_kwh": "C20",   # Annual electric usage (kWh)
    "annual_fuel_usage": "C21",     # Annual fuel usage

    # Control System (rows 30-35)
    "project_type": "C30",          # Dropdown: Proposed Control activity
    "demand_response_curtailment": "C31",  # Dropdown: Yes/No
    "bms_manufacturer": "B33",      # BMS Manufacturer
    "bms_product_type": "C33",      # Product type
    "total_project_cost": "C34",    # Total Proposed Project Cost
    "notes": "C35",                 # Notes

    # Subscription (rows 38-44)
    "subscription_product": "C38",
    "subscription_first_year_hardware": "C39",
    "subscription_previous_incentive": "C40",
    "subscription_years": "C41",
    "subscription_install_cost": "C42",
    "subscription_annual_fee": "C43",
    "subscription_notes": "C44",
}

# Affected Areas: columns C-G for areas 1-5
AREA_COLUMNS = {1: "C", 2: "D", 3: "E", 4: "F", 5: "G"}

AREA_FIELD_ROWS = {
    "project_affected_sqft": 49,
    "area_description": 50,
    "area_occupancy": 51,
    "area_project_type": 52,
    "is_new_equipment": 54,
    "ventilation_type": 55,
    "primary_heating": 56,
    "primary_cooling": 57,
    "terminal_units": 58,
    "secondary_heating_to_hp": 59,
    # Sequences of Operation (1 or 0)
    "seq_system_schedules": 61,
    "seq_optimal_start_stop": 62,
    "seq_reset_chilled_water": 63,
    "seq_reset_static_pressure": 64,
    "seq_reset_boiler_water": 65,
    "seq_demand_control_ventilation": 66,
    "seq_economizer_control": 67,
    "seq_reset_supply_air_temp": 68,
    "seq_reset_condenser_water": 69,
    # Optimization (subscription only)
    "opt_cooling": 71,
    "opt_ventilation": 72,
    "opt_heating": 73,
}


def _resolve_value(value):
    """Convert enum values to their string representation."""
    if isinstance(value, Enum):
        return value.value
    return value


TEMPLATES_DIR = Path(__file__).parent / "templates"


def _find_template(template_path=None):
    """Locate the BMS Calculator template file.

    Search order:
      1. Explicit path (if provided)
      2. Bundled templates/ directory (primary — ships with the project)
      3. Uploads folder (fallback for dev/testing)
    """
    search_paths = []
    if template_path:
        search_paths.append(Path(template_path))

    # Bundled template directory is the canonical location
    search_paths.append(TEMPLATES_DIR / TEMPLATE_FILENAME)

    # Also check for any .xlsx in templates/ (in case admin renamed)
    if TEMPLATES_DIR.exists():
        for f in sorted(TEMPLATES_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
            if f not in search_paths:
                search_paths.append(f)

    # Fallback locations
    base = Path(__file__).parent
    search_paths.extend([
        base / TEMPLATE_FILENAME,
        Path("/sessions/stoic-compassionate-ride/mnt/uploads") / TEMPLATE_FILENAME,
        base.parent / TEMPLATE_FILENAME,
    ])

    for p in search_paths:
        if p.exists():
            return str(p)
    return None


def get_template_info():
    """Return info about the currently active template (for admin UI)."""
    tmpl = _find_template()
    if not tmpl:
        return {"has_template": False}
    p = Path(tmpl)
    return {
        "has_template": True,
        "filename": p.name,
        "path": str(p),
        "size_bytes": p.stat().st_size,
        "modified": p.stat().st_mtime,
    }


def update_template(new_template_bytes: bytes, filename: str = None) -> dict:
    """Replace the bundled BMS Calculator template (admin operation).

    Args:
        new_template_bytes: Raw bytes of the uploaded .xlsx file
        filename: Original filename (preserved for reference)

    Returns:
        dict with success status and info
    """
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

    # Archive old template if it exists
    target = TEMPLATES_DIR / (filename or TEMPLATE_FILENAME)
    old_tmpl = _find_template()
    if old_tmpl and Path(old_tmpl).exists() and Path(old_tmpl).parent == TEMPLATES_DIR:
        archive_dir = TEMPLATES_DIR / "archive"
        archive_dir.mkdir(exist_ok=True)
        import time
        ts = time.strftime("%Y%m%d_%H%M%S")
        archive_name = f"{Path(old_tmpl).stem}_{ts}{Path(old_tmpl).suffix}"
        shutil.move(old_tmpl, str(archive_dir / archive_name))

    # Write new template
    with open(target, "wb") as f:
        f.write(new_template_bytes)
    os.chmod(str(target), 0o644)

    return {
        "success": True,
        "filename": target.name,
        "size_bytes": len(new_template_bytes),
        "archived_previous": old_tmpl is not None,
    }


def generate_calculator(
    calculator_input: BMSCalculatorInput,
    output_path: str,
    template_path: str = None,
) -> GenerationResponse:
    """
    Generate a completed Prescriptive BMS Calculator Excel file.
    Populates the actual template if available; falls back to standalone sheet.
    """
    warnings = []

    try:
        tmpl = _find_template(template_path)

        if tmpl:
            # Copy template to output path and ensure it's writable
            shutil.copy2(tmpl, output_path)
            os.chmod(output_path, 0o644)
            wb = openpyxl.load_workbook(output_path)
            _populate_template(wb, calculator_input, warnings)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = SHEET_USER_INPUTS
            _create_standalone_sheet(ws, calculator_input, warnings)
            warnings.append("Template not found — generated standalone data export. "
                          "For full calculator functionality, provide the official template.")

        wb.save(output_path)

        incentive = _estimate_incentive(calculator_input)
        if incentive:
            warnings.append(f"Estimated incentive: ${incentive:,.2f} (subject to PA review)")

        return GenerationResponse(
            success=True,
            file_path=output_path,
            incentive_estimate=incentive,
            warnings=warnings,
        )

    except Exception as e:
        return GenerationResponse(
            success=False,
            errors=[f"Excel generation failed: {str(e)}"],
        )


def _populate_template(wb, data: BMSCalculatorInput, warnings: list):
    """Populate the actual BMS Calculator template with form data."""

    # 1. Populate PA Only - Pre-Inspection sheet
    if SHEET_PRE_INSPECTION in wb.sheetnames:
        ws_pre = wb[SHEET_PRE_INSPECTION]
        for field, cell_ref in PRE_INSPECTION_MAP.items():
            value = getattr(data, field, None)
            if value is not None:
                ws_pre[cell_ref] = _resolve_value(value)
    else:
        warnings.append("Pre-Inspection sheet not found; populating User Inputs directly.")

    # 2. Populate User Inputs and Savings sheet
    if SHEET_USER_INPUTS not in wb.sheetnames:
        warnings.append("'User Inputs and Savings' sheet not found in template.")
        return

    ws = wb[SHEET_USER_INPUTS]

    # Direct input fields
    for field, cell_ref in USER_INPUTS_MAP.items():
        value = getattr(data, field, None)
        if value is not None:
            ws[cell_ref] = _resolve_value(value)

    # If Pre-Inspection sheet wasn't available, set header fields directly
    if SHEET_PRE_INSPECTION not in wb.sheetnames:
        direct_header = {
            "company_name": "C4",
            "company_address": "C5",
            "company_city": "F5",
            "customer_contact_name": "C6",
            "customer_phone": "F6",
            "electric_account": "C9",
            "gas_account": "C10",
            "electric_pa": "I6",
            "gas_pa": "I5",
        }
        for field, cell_ref in direct_header.items():
            value = getattr(data, field, None)
            if value is not None:
                ws[cell_ref] = _resolve_value(value)

    # 3. Populate Affected Areas
    for area in data.affected_areas:
        col = AREA_COLUMNS.get(area.area_number)
        if not col:
            continue

        for field, row in AREA_FIELD_ROWS.items():
            value = getattr(area, field, None)
            if value is not None:
                cell_ref = f"{col}{row}"
                ws[cell_ref] = _resolve_value(value)


def _create_standalone_sheet(ws, data: BMSCalculatorInput, warnings: list):
    """Create a standalone sheet with all calculator data (fallback when no template)."""

    header_font = Font(name="Arial", size=14, bold=True)
    section_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    label_font = Font(name="Arial", size=10)
    value_font = Font(name="Arial", size=10, bold=True)
    input_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")

    ws.merge_cells("A1:G1")
    ws["A1"] = "BMS Boss — Prescriptive BMS Calculator Data Export"
    ws["A1"].font = header_font

    ws.merge_cells("A2:G2")
    ws["A2"] = "Generated data for Mass Save Prescriptive BMS Calculator (2026 V1.0)"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)

    row = 4

    row = _write_section_header(ws, row, "Company / Account Information", section_font, section_fill)
    company_fields = [
        ("Company Name", data.company_name),
        ("Company Address", data.company_address),
        ("City", data.company_city),
        ("Customer Contact", data.customer_contact_name),
        ("Phone", data.customer_phone),
        ("PA Technical Rep", data.pa_technical_rep),
        ("Application #", data.application_number),
        ("Electric Account", data.electric_account),
        ("Gas Account", data.gas_account),
        ("Electric PA", data.electric_pa),
        ("Gas PA", data.gas_pa),
    ]
    row = _write_field_rows(ws, row, company_fields, label_font, value_font, input_fill)
    row += 1

    row = _write_section_header(ws, row, "Building Energy Use Intensity", section_font, section_fill)
    building_fields = [
        ("Principal Building Activity", _resolve_value(data.building_activity)),
        ("Non-Electric Heating Fuel", _resolve_value(data.heating_fuel)),
        ("Total Building Area (sqft)", data.total_building_sqft),
        ("Annual Electric Usage (kWh)", data.annual_electric_kwh),
        ("Annual Fuel Usage", data.annual_fuel_usage),
    ]
    row = _write_field_rows(ws, row, building_fields, label_font, value_font, input_fill)
    row += 1

    row = _write_section_header(ws, row, "Control System Information", section_font, section_fill)
    control_fields = [
        ("Proposed Control Activity", _resolve_value(data.project_type)),
        ("Demand Response Curtailment", data.demand_response_curtailment),
        ("BMS Manufacturer & Product", data.bms_manufacturer),
        ("Total Proposed Project Cost", f"${data.total_project_cost:,.2f}" if data.total_project_cost else None),
        ("Notes", data.notes),
    ]
    row = _write_field_rows(ws, row, control_fields, label_font, value_font, input_fill)
    row += 1

    row = _write_section_header(ws, row, "Affected Areas & Sequences of Operation", section_font, section_fill)

    if data.affected_areas:
        headers = ["Field"] + [f"Area {a.area_number}" for a in data.affected_areas]
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=row, column=col_idx + 1, value=header)
            cell.font = Font(name="Arial", size=10, bold=True)
        row += 1

        area_fields = [
            ("project_affected_sqft", "Affected Area (sqft)"),
            ("area_description", "Description"),
            ("ventilation_type", "Ventilation System"),
            ("primary_heating", "Primary Heating"),
            ("primary_cooling", "Primary Cooling"),
            ("terminal_units", "Terminal Units"),
            ("seq_system_schedules", "Seq: System Schedules"),
            ("seq_optimal_start_stop", "Seq: Optimal Start/Stop"),
            ("seq_reset_chilled_water", "Seq: Reset Chilled Water Temp"),
            ("seq_reset_static_pressure", "Seq: Reset Static Pressure"),
            ("seq_reset_boiler_water", "Seq: Reset Boiler Water Temp"),
            ("seq_demand_control_ventilation", "Seq: Demand Control Ventilation"),
            ("seq_economizer_control", "Seq: Economizer Control"),
            ("seq_reset_supply_air_temp", "Seq: Reset Supply Air Temp"),
            ("seq_reset_condenser_water", "Seq: Reset Condenser Water Temp"),
        ]

        for field_name, label in area_fields:
            ws.cell(row=row, column=1, value=label).font = label_font
            for col_idx, area in enumerate(data.affected_areas):
                value = getattr(area, field_name, None)
                if value is not None:
                    display = _resolve_value(value)
                    cell = ws.cell(row=row, column=col_idx + 2, value=display)
                    cell.font = value_font
                    cell.fill = input_fill
            row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 20

    warnings.append("Standalone data export generated (template not available).")


def _write_section_header(ws, row, title, font, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = font
    cell.fill = fill
    return row + 1


def _write_field_rows(ws, row, fields, label_font, value_font, input_fill):
    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = label_font
        if value is not None:
            cell = ws.cell(row=row, column=2, value=value)
            cell.font = value_font
            cell.fill = input_fill
        row += 1
    return row


def _estimate_incentive(data: BMSCalculatorInput) -> float:
    """
    Estimate the incentive: Rate x Sequence Count x Affected Area.
    Capped at 60% of total project costs.
    """
    if not data.affected_areas:
        return 0.0

    rates = {
        "Installation of New BMS": 0.10,
        "Add-On or Optimization of Sequences on Existing BMS": 0.05,
        "Subscription Based Control": 0.01,
    }

    rate = 0.0
    if data.project_type:
        pt_val = data.project_type.value if hasattr(data.project_type, 'value') else data.project_type
        rate = rates.get(pt_val, 0.0)

    total_incentive = 0.0
    for area in data.affected_areas:
        sqft = area.project_affected_sqft or 0
        sequences = sum([
            area.seq_system_schedules, area.seq_optimal_start_stop,
            area.seq_reset_chilled_water, area.seq_reset_static_pressure,
            area.seq_reset_boiler_water, area.seq_demand_control_ventilation,
            area.seq_economizer_control, area.seq_reset_supply_air_temp,
            area.seq_reset_condenser_water,
        ])
        total_incentive += rate * sequences * sqft

    if data.total_project_cost and data.total_project_cost > 0:
        cap = data.total_project_cost * 0.60
        total_incentive = min(total_incentive, cap)

    return total_incentive


def merge_bill_data_to_calculator(
    bill_data: ExtractedBillData,
    calculator_input: BMSCalculatorInput
) -> BMSCalculatorInput:
    """
    Merge extracted bill data into calculator input fields.
    Bill data populates auto-fillable fields; manual fields are preserved.
    """
    if bill_data.customer_name and not calculator_input.company_name:
        calculator_input.company_name = bill_data.customer_name

    if bill_data.service_address and not calculator_input.company_address:
        calculator_input.company_address = bill_data.service_address

    if bill_data.service_city and not calculator_input.company_city:
        calculator_input.company_city = bill_data.service_city

    if bill_data.account_number and not calculator_input.electric_account:
        calculator_input.electric_account = bill_data.account_number

    if bill_data.annual_usage_kwh and not calculator_input.annual_electric_kwh:
        calculator_input.annual_electric_kwh = bill_data.annual_usage_kwh

    # Set the electric PA based on the bill sponsor
    if bill_data.utility_sponsor:
        sponsor_val = bill_data.utility_sponsor.value if hasattr(bill_data.utility_sponsor, 'value') else bill_data.utility_sponsor
        calculator_input.electric_pa = sponsor_val

    return calculator_input
